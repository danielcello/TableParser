from bs4 import BeautifulSoup, BeautifulStoneSoup
from openpyxl import Workbook
import os

#needed variables
place = os.getcwd()
try:
  os.mkdir('output')
except:
  pass
os.chdir(os.path.join(os.getcwd(), 'output'))
title = input('name of file?\n')


#parser
def arrayCreate(table):
  sparcer = table.find_all('tr')
  print(sparcer)
  insertTable = []
  for x in range(len(sparcer)):
    insertSection = []
    if not (sparcer[x].find_all('th')):
      section = sparcer[x].find_all('td')
    else:
      section = sparcer[x].find_all('th')
    for b in range(len(section)):
      try:
        insertSection.append(section[b].get_text())
      except:
        pass
    insertTable.append(insertSection)
    print(insertTable)

  return insertTable


table = input("Please enter your table: \n")

htmlParse = BeautifulSoup(table, "html.parser")
print(htmlParse)
print("------------------------\ntable-ing...\n------------------------")
table = htmlParse.find('table')
print(table)
print(
  "------------------------\nnow let's seperate it...\n------------------------\n"
)
arrayItem = arrayCreate(table)

# going to excel
wb = Workbook()
ws = wb.active
ws.title = title
print(
  "------------------------\nturning it into an excel sheet...\n------------------------\n"
)
for i in range(len(arrayItem)):
  ws.append(arrayItem[i])
  print(ws)
#  for x in range(len(arrayItem[i])):
#    print(arrayItem[i][x])
#    the_cell = ws.cell(row=(i + 1), column=(x + 1))
#    print(the_cell)
#    the_cell.value = arrayItem[i][x]
print(ws['A1'].value)
wb.save(f'{title}.xlsx')
